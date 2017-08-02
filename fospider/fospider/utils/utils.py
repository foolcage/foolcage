import datetime
import json
import logging
import os

import openpyxl

from fospider import settings
from fospider.items import SecurityItem

logger = logging.getLogger(__name__)


def init_log():
    root_logger = logging.getLogger()
    root_logger.setLevel(logging.DEBUG)

    fh = logging.FileHandler('foolcage.log')
    fh.setLevel(logging.DEBUG)

    ch = logging.StreamHandler()
    ch.setLevel(logging.DEBUG)

    # create formatter and add it to the handlers
    formatter = logging.Formatter("%(levelname) -10s %(asctime)s %(module)s:%(lineno)s %(funcName)s %(message)s")
    fh.setFormatter(formatter)
    ch.setFormatter(formatter)

    # add the handlers to the logger
    root_logger.addHandler(fh)
    root_logger.addHandler(ch)


def chrome_copy_header_to_dict(src):
    lines = src.split('\n')
    header = {}
    if lines:
        for line in lines:
            try:
                index = line.index(':')
                key = line[:index]
                value = line[index + 1:]
                if key and value:
                    header.setdefault(key.strip(), value.strip())
            except Exception:
                pass
    return header


def is_sh_stock_file(path):
    return path.endswith(settings.SH_STOCK_FILE)


def is_sz_stock_file(path):
    return path.endswith(settings.SZ_STOCK_FILE)


def get_security_item(path):
    if is_sh_stock_file(path):
        return get_sh_security_item(path)
    elif is_sz_stock_file(path):
        return get_sz_security_item(path)


def generate_csv_line(*items):
    if items:
        result = items[0]
        for item in items[1:]:
            result = (result + ',' + item)
        return result
    return ''


def get_sz_security_item(path):
    wb = openpyxl.load_workbook(path)
    for name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(name)
        max_row, max_column = sheet.max_row, sheet.max_column
        for i in range(2, max_row):
            code = sheet.cell(row=i, column=1).value
            name = sheet.cell(row=i, column=2).value
            list_date = sheet.cell(row=i, column=8).value
            # ignore just in B
            if not list_date:
                continue
            yield SecurityItem(id=gen_security_id('stock', 'sz', code), type='stock', exchange='sz', code=code,
                               name=name, listDate=list_date)


def gen_security_id(type, exchange, code):
    return type + '_' + exchange + '_' + code;


def get_sh_security_item(path):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    with open(path, encoding=encoding) as fr:
        lines = fr.readlines()
        for line in lines[1:]:
            code, name, _, _, list_date, _, _ = line.split()
            yield SecurityItem(id=gen_security_id('stock', 'sh', code), type='stock', exchange='sh', code=code,
                               name=name, listDate=list_date)


def get_tick_item(path, the_date, security_item):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    with open(path, encoding=encoding) as fr:
        lines = fr.readlines()
        for line in lines[1:]:
            tmp_timestamp, price, tmp_change, volume, turnover, tmp_direction = line.split()
            # timestamp = datetime.datetime.strptime(the_date + tmp_timestamp, '%Y-%m-%d%H:%M:%S')
            timestamp = the_date + " " + tmp_timestamp
            change = 0
            if not tmp_change == '--':
                change = float(tmp_change)
            direction = 0
            if tmp_direction == '买盘':
                direction = 1
            elif tmp_direction == '卖盘':
                direction = -1

            # yield SecurityItem(code_id='sh' + code, code=code, name=name, list_date=list_date, exchange='sh',
            #                    type='stock')
            # yield generate_csv_line(code, name, list_date, 'sh', 'stock')
            yield {"securityId": security_item['id'],
                   "code": security_item['code'],
                   "timestamp": timestamp,
                   "price": price,
                   "change": change,
                   "direction": direction,
                   "volume": volume,
                   "turnover": turnover}


def detect_encoding(url):
    import urllib.request
    from chardet.universaldetector import UniversalDetector

    usock = urllib.request.urlopen(url)
    detector = UniversalDetector()
    for line in usock.readlines():
        detector.feed(line)
        if detector.done: break
    detector.close()
    usock.close()
    return detector.result.get('encoding')


def setup_env():
    if not os.path.exists('data'):
        os.makedirs('data')
    pass
    # db_setup()


def mkdir_for_security(item):
    fuquan_kdata_dir = get_fuquan_kdata_dir(item)
    if not os.path.exists(fuquan_kdata_dir):
        os.makedirs(fuquan_kdata_dir)

    finance_dir = get_finance_dir(item)
    if not os.path.exists(finance_dir):
        os.makedirs(finance_dir)


def get_security_dir(item):
    return os.path.join(settings.FILES_STORE, item['type'], item['exchange'], item['code'])


def get_kdata_dir(item):
    return os.path.join(get_security_dir(item), 'kdata')


def get_fuquan_kdata_dir(item):
    return os.path.join(get_kdata_dir(item), 'fuquan')


def get_kdata_path(item, year, quarter, fuquan):
    if fuquan:
        return os.path.join(get_fuquan_kdata_dir(item), '{}_{}_fuquan_dayk.json'.format(year, quarter))
    else:
        return os.path.join(get_kdata_dir(item), '{}_{}_dayk.json'.format(year, quarter))


def get_trading_dates_path(item):
    return os.path.join(get_security_dir(item), 'trading_dates.json')


def get_trading_dates(item):
    dates = []
    dates_path = get_trading_dates_path(item)
    try:
        with open(dates_path) as data_file:
            dates = json.load(data_file)
    except Exception as e:
        logger.info(e)

    if not dates:
        dir = get_kdata_dir(item)
        files = [os.path.join(dir, f) for f in os.listdir(dir) if os.path.isfile(os.path.join(dir, f))]

        for f in files:
            with open(f) as data_file:
                items = json.load(data_file)
                for item in items:
                    dates.append(item['date'])
    dates.sort()
    return dates


def get_tick_dir(item):
    return os.path.join(settings.FILES_STORE, item['type'], item['exchange'], item['code'], 'tick')


def get_tick_path(item, date):
    return os.path.join(get_tick_dir(item), date + ".xls")


def get_finance_dir(item):
    return os.path.join(get_security_dir(item), "finance")


def get_balance_sheet_path(item):
    return os.path.join(get_finance_dir(item), "balance_sheet.xls")


def get_income_statement_path(item):
    return os.path.join(get_finance_dir(item), "income_statement.xls")


def get_cash_flow_statement_path(item):
    return os.path.join(get_finance_dir(item), "cash_flow_statement.xls")


def get_sh_stock_list_path():
    return os.path.join(settings.FILES_STORE, settings.SH_STOCK_FILE)


def get_sz_stock_list_path():
    return os.path.join(settings.FILES_STORE, settings.SZ_STOCK_FILE)


def is_available_tick(path):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    try:
        with open(path, encoding=encoding) as fr:
            line = fr.readline()
            return u'成交时间', u'成交价', u'价格变动', u'成交量(手)', u'成交额(元)', u'性质' == line.split()
    except Exception:
        return False


# time utils
def get_datetime(str):
    return datetime.datetime.strptime(str, "%Y-%m-%d")


def get_year_quarter(time):
    return time.year, (time.month // 3) + 1


def get_quarters(start):
    start_time = get_datetime(start)
    today = datetime.date.today()
    start_year_quarter = get_year_quarter(start_time)
    current_year_quarter = get_year_quarter(today)
    if current_year_quarter[0] == start_year_quarter[0]:
        return [(current_year_quarter[0], x) for x in range(start_year_quarter[1], current_year_quarter[1] + 1)]
    elif current_year_quarter[0] - start_year_quarter[0] == 1:
        return [(start_year_quarter[0], x) for x in range(start_year_quarter[1], 5)] + \
               [(current_year_quarter[0], x) for x in range(1, current_year_quarter[1] + 1)]
    elif current_year_quarter[0] - start_year_quarter[0] > 1:
        return [(start_year_quarter[0], x) for x in range(start_year_quarter[1], 5)] + \
               [(x, y) for x in range(start_year_quarter[0] + 1, current_year_quarter[0]) for y in range(1, 5)] + \
               [(current_year_quarter[0], x) for x in range(1, current_year_quarter[1] + 1)]
    else:
        raise Exception("wrong start time:{}".format(start));
