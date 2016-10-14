import json
import os

import openpyxl
import rethinkdb as r

from fospider import settings
from fospider.items import SecurityItem


def chrome_copy_header_to_dict(src):
    lines = src.split('\n')
    header = {}
    if lines:
        for line in lines:
            try:
                index = line.index(':')
                key = line[:index]
                value = line[index + 1:]
                if key and value:
                    header.setdefault(key.strip(), value.strip())
            except Exception:
                pass
    return header


def get_security_item(path):
    if path.endswith("sh.txt"):
        return get_sh_security_item(path)
    elif path.endswith("sz.xlsx"):
        return get_sz_security_item(path)


def get_tick_item(path):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    with open(path, encoding=encoding) as fr:
        lines = fr.readlines()
        for line in lines[1:]:
            code, name, _, _, list_date, _, _ = line.split()
            yield SecurityItem(code='sh' + code, name=name, list_date=list_date, exchange='sh', type='stock')


def get_sz_security_item(path):
    wb = openpyxl.load_workbook(path)
    for name in wb.get_sheet_names():
        sheet = wb.get_sheet_by_name(name)
        max_row, max_column = sheet.max_row, sheet.max_column
        for i in range(2, max_row):
            code = sheet.cell(row=i, column=1).value
            name = sheet.cell(row=i, column=2).value
            list_date = sheet.cell(row=i, column=8).value
            yield SecurityItem(code='sz' + code, name=name, list_date=list_date, exchange='sz', type='stock')


def get_sh_security_item(path):
    encoding = settings.DOWNLOAD_TXT_ENCODING if settings.DOWNLOAD_TXT_ENCODING else detect_encoding(
        url='file://' + os.path.abspath(path)).get('encoding')
    with open(path, encoding=encoding) as fr:
        lines = fr.readlines()
        for line in lines[1:]:
            code, name, _, _, list_date, _, _ = line.split()
            yield SecurityItem(code=code, name=name, list_date=list_date, exchange='sh', type='stock')


def detect_encoding(url):
    import urllib.request
    from chardet.universaldetector import UniversalDetector

    usock = urllib.request.urlopen(url)
    detector = UniversalDetector()
    for line in usock.readlines():
        detector.feed(line)
        if detector.done: break
    detector.close()
    usock.close()
    return detector.result.get('encoding')


# database info
RDB_HOST = os.environ.get('RDB_HOST') or 'localhost'
RDB_PORT = os.environ.get('RDB_PORT') or 28015

FOOLCAGE_DB = 'foolcage'

TABLE_EXCHANGE = 'exchange'
TABLE_SECURITY_TYPE = 'security_type'
TABLE_SECURITY = 'security'
TABLE_TICK = 'tick'

conn = None


def create_tables(conn):
    if not r.table_list().contains(TABLE_EXCHANGE):
        r.db(FOOLCAGE_DB).table_create(TABLE_EXCHANGE).run(conn)
    if not r.table_list().contains(TABLE_SECURITY_TYPE):
        r.db(FOOLCAGE_DB).table_create(TABLE_SECURITY_TYPE).run(conn)
    if not r.table_list().contains(TABLE_SECURITY):
        r.db(FOOLCAGE_DB).table_create(TABLE_SECURITY).run(conn)
    if not r.table_list().contains(TABLE_TICK):
        r.db(FOOLCAGE_DB).table_create(TABLE_TICK).run(conn)


def db_setup():
    try:
        conn = r.connect(host=RDB_HOST, port=RDB_PORT)

        if not r.db_list().contains(FOOLCAGE_DB):
            r.db_create(FOOLCAGE_DB).run(conn)
        create_tables(conn)
    except r.ReqlDriverError or r.ReqlDriverError as error:
        print(error.message)


def db_get_securities():
    selection = list(r.table(TABLE_SECURITY).run(conn))
    return json.dumps(selection)


def db_insert_security(item):
    try:
        r.db(FOOLCAGE_DB).table(TABLE_SECURITY).insert(item).run(conn)
    except r.RqlRuntimeError as err:
        print(err.message)
